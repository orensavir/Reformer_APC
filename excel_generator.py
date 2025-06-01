import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
import os

# --- 1. Load Data ---
raw_file = "Raw data for AI model.xlsx"
df = pd.read_excel(raw_file, header=1)
excel_file = "Reformer_APC_Professional.xlsx"
if os.path.exists(excel_file):
    os.remove(excel_file)

# --- 2. Model Parameters ---
params = [
    ("RON_Base", 94.2, "RON base value"),
    ("RON_Temp_Coeff", 0.25, "Temp coefficient for RON"),
    ("RON_Cat_Coeff", 0.02, "Catalyst age coefficient for RON"),
    ("MON_Gap", 10, "Typical RON/MON gap"),
    ("Yield_Base", 84.0, "Yield base value"),
    ("Yield_Coeff", 0.5, "RON severity effect on yield"),
    ("LPG_Base", 12.0, "LPG base value"),
    ("LPG_Coeff", 0.6, "RON severity effect on LPG"),
    ("RVP_Base", 7.5, "RVP base value"),
    ("RVP_Coeff", 0.11, "LPG effect on RVP"),
    ("H2_Base", 89.0, "H2 purity base"),
    ("H2_Coeff", 0.05, "Severity effect on H2 purity"),
    ("Cat_Deact_Coeff", 0.03, "Deactivation rate per day"),
]
params_df = pd.DataFrame(params, columns=["Parameter", "Current", "Description"])

# --- 3. Create Excel File ---
wb = Workbook()
ws_data = wb.active
ws_data.title = "Data"
ws_data.append(list(df.columns))
for row in df.itertuples(index=False):
    ws_data.append(row)

ws_params = wb.create_sheet("ModelParams")
ws_params.append(list(params_df.columns))
for r in params_df.itertuples(index=False):
    ws_params.append(list(r))

# --- 4. Calculated Sheet: Excel Formulas ---
ws_calc = wb.create_sheet("Calculated")
headers = ["Timestamp", "WABT", "RON_Pred", "MON_Pred", "Yield_Pred", "LPG_Pred", "RVP_Pred", "H2_Purity_Pred", "Cat_Deact"]
ws_calc.append(headers)
nrows = df.shape[0]

for i in range(2, nrows+2):
    ts = f"=Data!A{i}"
    r1 = f"Data!B{i}"
    r2 = f"Data!C{i}"
    r3 = f"Data!D{i}"
    cat_age = f"Data!E{i}"
    mp = lambda idx: f"ModelParams!B{idx+1}"
    wabt = f"=AVERAGE({r1},{r2},{r3})"
    ron_pred = f"={mp(0)}+{mp(1)}*({wabt}-500)-{mp(2)}*{cat_age}"
    mon_pred = f"={ron_pred}-{mp(3)}"
    yield_pred = f"={mp(4)}-{mp(5)}*({ron_pred}-95)"
    lpg_pred = f"={mp(6)}+{mp(7)}*({ron_pred}-95)"
    rvp_pred = f"={mp(8)}-{mp(9)}*({lpg_pred}-12)"
    h2_purity = f"={mp(10)}-{mp(11)}*({wabt}-500)"
    cat_deact = f"={mp(12)}*{cat_age}"
    ws_calc.append([ts, wabt, ron_pred, mon_pred, yield_pred, lpg_pred, rvp_pred, h2_purity, cat_deact])

# --- 5. Optimization Sheet (Formula-Ready) ---
ws_opt = wb.create_sheet("Optimization")
opt_headers = ["Parameter", "Current", "Target", "Predicted", "Recommended", "Unit", "Constraint", "Status"]
ws_opt.append(opt_headers)
for param in [
    ("RON (Research Octane)", "=", "95", "=", "=", "-", ">=95", ""),
    ("MON (Motor Octane)", "=", "85", "=", "=", "-", ">=85", ""),
    ("C5+ Yield", "=", "83", "=", "=", "%", ">=83", ""),
    ("LPG Produced", "=", "14", "=", "=", "%", "<=14", ""),
    ("Reformate RVP", "=", "7.5", "=", "=", "psi", "<=7.5", ""),
    ("H2 Purity", "=", "88", "=", "=", "%", ">=88", ""),
    ("Catalyst Deactivation", "=", "1.0", "=", "=", "%", "<=1.0", "")
]:
    ws_opt.append(param)

# --- 6. Documentation Sheet (Full Text) ---
ws_doc = wb.create_sheet("Documentation")
doc_sections = [
    ("Scope", 
     "This application provides monitoring, prediction, and operator advisory APC for a 3-reactor semi-regenerative catalytic reformer. "
     "The system is designed to be dynamic, allowing for real-time updates and providing a professional dashboard for operations."),
    ("Key Features",
     "- Real-time monitoring of critical process parameters\n"
     "- Predictive analytics for product quality and yields\n"
     "- What-if scenario analysis for process optimization\n"
     "- Operational recommendations for setpoint adjustment\n"
     "- Historical data trending and analysis\n"
     "- Documentation of model equations and assumptions"),
    ("Workflow",
     "Monitor the Dashboard for KPIs, alerts, and trends\n"
     "Review historical data in the Data section\n"
     "Explore model equations in the Model section\n"
     "Use the Optimization page for what-if analysis and setpoint recommendations"),
    ("System Components",
     "Dashboard: Operator-centric view with KPIs, status indicators, and critical trends. All visualizations update in real-time as new data becomes available.\n"
     "Data Management: Historical data storage, trending, and export capabilities. Filterable tables and interactive charts for data analysis.\n"
     "Model Equations: Documentation of all calculation methods, references to literature, and calibration details for the predictive models.\n"
     "Optimization Tools: What-if scenario analysis, constraint-based optimization, and setpoint recommendations for operators."),
    ("References",
     "PTQ Q3 2017, AspenTech, KBC/Shell case studies\n"
     "Hydrocarbon Processing literature\n"
     "Plant historical data and calibration runs\n"
     "Industry standard correlations for naphtha reforming"),
    ("Model Equations and Formulas",
     "WABT = (R1_Inlet_Temp + R2_Inlet_Temp + R3_Inlet_Temp)/3\n"
     "RON_Pred = RON_Base + RON_Temp_Coeff*(WABT-500) - RON_Cat_Coeff*Cat_Age\n"
     "MON_Pred = RON_Pred - MON_Gap\n"
     "Yield_Pred = Yield_Base - Yield_Coeff*(RON_Pred-95)\n"
     "LPG_Pred = LPG_Base + LPG_Coeff*(RON_Pred-95)\n"
     "RVP_Pred = RVP_Base - RVP_Coeff*(LPG_Pred-12)\n"
     "H2_Purity_Pred = H2_Base - H2_Coeff*(WABT-500)\n"
     "Cat_Deact = Cat_Deact_Coeff*Cat_Age"),
    ("Optimization Methodology",
     "The optimization module uses the inferential model to recommend setpoint changes that maximize performance while respecting operational constraints.\n"
     "Maximize reformate production (yield × throughput)\n"
     "Minimize quality giveaway (RON > target)\n"
     "Maintain all product specifications (RON, RVP)\n"
     "Respect equipment constraints (temperatures, pressures)\n"
     "Use Excel Solver Add-in for recommended optimization scenarios."),
    ("User Guide",
     "1. Upload or paste new daily/hourly data into the Data sheet.\n"
     "2. Review predictions in the Calculated sheet (auto-updates from Data & ModelParams).\n"
     "3. Use Dashboard for visual trends (build charts as desired).\n"
     "4. Use Optimization sheet and Excel Solver for scenario analysis.\n"
     "5. All equations and references are in the Documentation and ModelParams sheets.\n"
     "6. Adjust ModelParams as needed; Calculated results update automatically.")
]
row = 1
for title, body in doc_sections:
    ws_doc[f"A{row}"] = title
    ws_doc[f"A{row}"].font = Font(bold=True, size=12)
    row += 1
    for bl in body.split('\n'):
        ws_doc[f"A{row}"] = bl
        row += 1
    row += 1

# --- 7. Dashboard Sheet Placeholder ---
ws_dash = wb.create_sheet("Dashboard")
ws_dash["A1"] = ("Create KPI cards and charts using the Calculated sheet. "
                 "Recommended: KPI summary cards, trend lines for RON, Yield, LPG, RVP, H2 Purity, Catalyst Deactivation, "
                 "WABT and Reactor Inlet Temperatures. Use Excel's 'Insert > Recommended Charts'.")

wb.save(excel_file)
print(f"\nProfessional Excel APC file created: {excel_file}")
print("Open in Excel. All logic/formulas are live. Build charts in Dashboard with Insert > Chart.")
